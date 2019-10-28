# import openpyxl
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook("sourav.xlsx")
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(1, 1)
# print(cell.value)
# Number Of ROW ******************
# print(sheet.max_row)
# for row in range(1, sheet.max_row + 1):
#       print(row)
for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)  # 3 Is Column Number ****************
    # print(f"Price Column Value {[row]} = {cell.value}")
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=5,
          max_col=5)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'f2')
# wb.save('sourav.xlsx')
wb.save('save.xlsx')
